from SDM.User_Interface.Sub_Windows.select_tab_window import SelectTabWindow
from tkinter import *
from tkinter import filedialog, StringVar, IntVar
from tkinter.filedialog import askopenfile
from tkinter import messagebox
import pandas as pd
import os
import openpyxl

class MergeFilesWindow(Toplevel):
  def __init__(self, parent):
    super().__init__(parent)
    self.title("Merge Files")
    self.geometry("500x700")
    self.parent = parent
    self.main = self.parent.parent
    self.files_to_merge = self.main.files_to_merge

    self.frame_options = Frame(self)
    self.frame_options.grid(row=0, column=0, padx=(70, 0), pady=3, sticky='nsew')

    self.data_to_merge = pd.DataFrame()
    self.merge_column_options = self.data_to_merge.columns.tolist() if len(self.data_to_merge.columns.tolist()) > 0 else ['SubID']

    self.mergeVariablesHeader = Label(self.frame_options, text = 'Settings: Merge other variables')
    self.mergeVariablesHeader.config(font=(None, 12, 'bold'))
    self.mergeVariablesHeader.grid(row=0, column=0, padx=5, pady=5)

    self.mergeVariablesNote = Label(self.frame_options, text = 'Note: Merge will use a subject ID as a key.\nSubject ID column required.')
    self.mergeVariablesNote.config(font=(None, 9, 'italic'))
    self.mergeVariablesNote.grid(row=1, column=0, padx=5, pady=5)

    self.merge_filename = ''
    self.selected_sheet = 0
    self.mergeExcelButton = Button(self.frame_options, width = 40, text='1. Select Excel file:', command=self.select_merge_file)
    self.mergeExcelButton.grid(row=2, column=0, padx=5, pady=(2, 5))

    self.mergeVariablesLabelText = '2. Select variables to merge:'
    self.mergeVariablesLabel = Label(self.frame_options, text = self.mergeVariablesLabelText)
    self.mergeVariablesListbox = Listbox(self.frame_options, selectmode=MULTIPLE,height=6)

    self.mergeSubidLabelText = '3. Select Subject ID column to use as merging key:'
    self.mergeSubidLabel = Label(self.frame_options, text = self.mergeSubidLabelText)
    self.subid_column = StringVar(self)
    self.subid_column.set('')
        # self.om_variable.trace('w', self.option_select)
    self.mergeSubidOptionMenu = OptionMenu(self.frame_options, self.subid_column, *self.merge_column_options)

    #display chosen files to merge
    self.addMergeFileButton = Button(self.frame_options, text='Submit file to merge with program results.', command=self.submit_data_to_merge, fg='blue')
    self.removeMergeFilesButton = Button(self.frame_options, text='Remove all merge files.', command=self.remove_data_to_merge, fg='red')

    self.frame_text = Frame(self)
    self.frame_text.grid(row=1, column=0, padx=(10, 0), pady=3, sticky='nsew')
    self.text_widget = Text(self.frame_text, wrap=WORD, width=55, height=15)
    self.scrollbar = Scrollbar(self.frame_text, command=self.text_widget.yview)
    self.text_widget.config(yscrollcommand=self.scrollbar.set)

    if len(self.files_to_merge.keys()) > 0:
      self.removeMergeFilesButton.grid(row=8, column = 0, padx=5, pady=4)
      if not self.text_widget.get("1.0", END).strip():
        self.text_widget.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
        self.scrollbar.grid(row=1, column=1, padx=(0, 5), pady=5, sticky="ns")
        self.text_widget.insert(END, "Submitted Merges:", "submittedMergesHeader")
        self.text_widget.tag_configure("submittedMergesHeader", font=(None, 10, "bold"))
      self.text_widget.delete("2.0", END)
      file_count = 1
      for filename, merge_file in self.files_to_merge.items():
        self.text_widget.insert(END, f'\n{merge_file["label"]}\n', f'label{file_count}')
        self.text_widget.insert(END, "-" * 55 + "\n", f'line{file_count}')
        self.text_widget.tag_configure(f'line{file_count}', foreground="gray")

    self.frame_options.grid_rowconfigure(0, weight=0)
    self.frame_options.grid_columnconfigure(0, weight=0)
    self.frame_options.grid_columnconfigure(1, weight=1)

  def select_merge_file(self):
    file = filedialog.askopenfile(mode='r', filetypes=[('Excel file to merge', '*.xlsx')])
    if file:
      excel_file = os.path.abspath(file.name)
      workbook = openpyxl.load_workbook(excel_file)
      self.selected_sheet == 0
      if len(workbook.sheetnames) > 1:
        select_tab_window = SelectTabWindow(self, workbook.sheetnames)
        select_tab_window.grab_set()
        self.wait_window(select_tab_window)
        self.lift(self.parent)
        self.selected_sheet = select_tab_window.selected_sheet
        print(self.selected_sheet)
      if len(workbook.sheetnames) > 1 and self.selected_sheet == 0:
        self.reset_merge_file_UI()
      else:
        self.data_to_merge = pd.read_excel(excel_file, sheet_name=self.selected_sheet)
        if len(self.data_to_merge) > 0:
          self.mergeExcelButton['text'] = f'1. Selected file: {str(file.name.split("/")[-1])}'
          self.merge_filename = file.name.split("/")[-1]
          self.mergeVariablesListbox.delete(0, 'end')
          menu = self.mergeSubidOptionMenu["menu"]
          menu.delete(0,'end')
          for col in self.data_to_merge.columns:
            self.mergeVariablesListbox.insert("end", col)
            menu.add_command(label=col, command=lambda value=col: self.subid_column.set(value))
          menu = self.mergeSubidOptionMenu["menu"]
          self.subid_column.set(self.data_to_merge.columns.tolist()[0])

          self.mergeVariablesLabel.grid(row=3, column=0, padx=0, pady=1)
          self.mergeVariablesListbox.grid(row=4, column=0, padx=5, pady=5)
          self.mergeSubidLabel.grid(row=5, column=0, padx=5, pady=0)
          self.mergeSubidOptionMenu.grid(row=6, column=0, padx=0, pady=7)
          self.addMergeFileButton.grid(row = 7, column = 0, padx=5, pady=(10, 0))
          self.removeMergeFilesButton.grid(row=8, column = 0, padx=5, pady=4)
        else:
          messagebox.showerror('SDM Error', 'Invalid file selected. Not able to load data.')

  def submit_data_to_merge(self):
    if len(self.data_to_merge) < 1:
      messagebox.showerror("Error", 'Invalid file chosen')
    elif len([self.mergeVariablesListbox.get(i) for i in self.mergeVariablesListbox.curselection()]) == 0:
      messagebox.showerror("Error", 'No variables chosen')
    elif (self.subid_column.get() not in self.data_to_merge.columns.tolist()) or (self.subid_column.get() in [self.mergeVariablesListbox.get(i) for i in self.mergeVariablesListbox.curselection()]):
      messagebox.showerror("Error", f'Invalid Subject ID column: {self.subid_column.get()}')
    else:
      self.files_to_merge[f'{self.merge_filename}_{self.selected_sheet}'] = {
        'df': self.data_to_merge.copy(),
        'variables': [self.mergeVariablesListbox.get(i) for i in self.mergeVariablesListbox.curselection()],
        'label': f'Filename: {self.merge_filename}\nSheet: {self.selected_sheet if self.selected_sheet else "1"}\nVariables: {", ".join([self.mergeVariablesListbox.get(i) for i in self.mergeVariablesListbox.curselection()])}\nMerge Key: {self.subid_column.get()}',
        'subid_column': self.subid_column.get()
      }
      if not self.text_widget.get("1.0", END).strip():
        self.text_widget.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
        self.scrollbar.grid(row=1, column=1, padx=(0, 5), pady=5, sticky="ns")
        self.text_widget.insert(END, "Submitted Merges:", "submittedMergesHeader")
        self.text_widget.tag_configure("submittedMergesHeader", font=(None, 10, "bold"))
      self.text_widget.delete("2.0", END)
      file_count = 1
      for filename, merge_file in self.files_to_merge.items():
        self.text_widget.insert(END, f'\n{merge_file["label"]}\n', f'label{file_count}')
        self.text_widget.insert(END, "-" * 55 + "\n", f'line{file_count}')
        self.text_widget.tag_configure(f'line{file_count}', foreground="gray")
        file_count += 1
        
      self.reset_merge_file_UI()

  def reset_merge_file_UI(self):
    self.data_to_merge = pd.DataFrame()
    self.merge_filename = ''
    self.mergeExcelButton['text'] = '1. Select Excel file to merge: '
    menu = self.mergeSubidOptionMenu["menu"]
    menu.delete(0,'end')
    self.subid_column.set('')
    menu.add_command(label='', command=lambda value='': self.subid_column.set(value))
    self.mergeVariablesListbox.delete(0, END)
    self.selected_sheet = 0

  def remove_data_to_merge(self):
    self.reset_merge_file_UI()
    self.text_widget.delete("1.0", END)
    self.text_widget.grid_forget()
    self.scrollbar.grid_forget()
    self.files_to_merge = {}
    self.main.update_files_to_merge(self.files_to_merge)

          
